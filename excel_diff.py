#!/usr/bin/env python3
"""
Excel Workbook Diff and Copy Tool

Compares two Excel workbooks across all tabs, highlights differences in a
copied output workbook, and generates a Diff_Summary sheet listing all
differences.

Usage:
  python excel_diff.py --base BASE.xlsx --new NEW.xlsx --out OUT.xlsx

Options:
  --include-sheets Sheet1,Sheet2  Compare only these sheets (comma-separated)
  --exclude-sheets Sheet3         Exclude these sheets (comma-separated)
  --tolerance 0.0                 Numeric tolerance for differences
  --case-insensitive              Compare text case-insensitively
  --trim-whitespace               Trim whitespace in text comparisons
  --no-highlights                 Do not highlight cells in output workbook
  --values                        Compare displayed values (default)
  --formulas                      Compare formulas instead of values

Requires: openpyxl
  pip install openpyxl

Notes:
  - The output workbook is a copy of the "new" workbook with differences
    highlighted by default and a Diff_Summary sheet added.
  - Added cells: present in new, empty/None in base (green fill).
  - Removed cells: present in base, empty/None in new (red fill).
  - Changed cells: present in both but different (yellow fill).
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from typing import Iterable, List, Optional, Set, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print("Missing dependency 'openpyxl'. Install with: pip install openpyxl", file=sys.stderr)
    raise


# Highlight fills
FILL_CHANGED = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # light yellow
FILL_ADDED = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # light green
FILL_REMOVED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red


@dataclass
class DiffRecord:
    sheet: str
    cell: str
    status: str  # 'added' | 'removed' | 'changed' | 'missing_sheet_new' | 'missing_sheet_base'
    old: Optional[object]
    new: Optional[object]


def is_number(x: object) -> bool:
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def normalize_str(s: Optional[str], *, trim_whitespace: bool, case_insensitive: bool) -> Optional[str]:
    if s is None:
        return None
    if not isinstance(s, str):
        return s  # pass-through non-strings
    out = s.strip() if trim_whitespace else s
    return out.lower() if case_insensitive else out


def equal_values(a: object, b: object, *, tolerance: float, trim_whitespace: bool, case_insensitive: bool) -> bool:
    # Exact equality shortcut
    if a == b:
        # Handles None, identical numbers/strings/dates
        return True

    # Numeric tolerance
    if is_number(a) and is_number(b):
        try:
            return abs(float(a) - float(b)) <= tolerance
        except Exception:
            return False

    # String normalization
    if isinstance(a, str) or isinstance(b, str):
        return normalize_str(str(a) if a is not None else None,
                             trim_whitespace=trim_whitespace,
                             case_insensitive=case_insensitive) == \
               normalize_str(str(b) if b is not None else None,
                             trim_whitespace=trim_whitespace,
                             case_insensitive=case_insensitive)

    # Fallback to inequality
    return False


def cell_coord(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def merged_non_anchors(ws: Worksheet) -> Set[Tuple[int, int]]:
    """Return set of coordinates (row, col) that are part of merged ranges but not the top-left anchor."""
    coords: Set[Tuple[int, int]] = set()
    for mr in getattr(ws, 'merged_cells', []):
        # mr is a MergedCellRange like 'A1:C3'
        min_row, min_col, max_row, max_col = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if r == min_row and c == min_col:
                    continue
                coords.add((r, c))
    return coords


def sheet_used_bounds(ws: Worksheet) -> Tuple[int, int]:
    """Heuristically get used bounds as (max_row, max_col)."""
    # openpyxl's calculate_dimension returns 'A1:D10' or 'A1'
    dim = ws.calculate_dimension()
    try:
        if ':' in dim:
            _, br = dim.split(':', 1)
        else:
            br = dim
        # extract trailing number and leading letters
        col_letters = ''.join([ch for ch in br if ch.isalpha()])
        row_digits = ''.join([ch for ch in br if ch.isdigit()])
        from openpyxl.utils.cell import column_index_from_string
        max_col = column_index_from_string(col_letters) if col_letters else 1
        max_row = int(row_digits or '1')
        return max_row, max_col
    except Exception:
        # Fallback
        return ws.max_row or 1, ws.max_column or 1


def iter_diff_cells(base_ws: Worksheet,
                    new_ws: Worksheet,
                    *,
                    tolerance: float,
                    trim_whitespace: bool,
                    case_insensitive: bool) -> Iterable[Tuple[int, int, str]]:
    """Yield (row, col, status) for cells that differ.
    status in {'added','removed','changed'}.
    """
    base_max_r, base_max_c = sheet_used_bounds(base_ws)
    new_max_r, new_max_c = sheet_used_bounds(new_ws)
    max_r = max(base_max_r, new_max_r)
    max_c = max(base_max_c, new_max_c)

    base_skip = merged_non_anchors(base_ws)
    new_skip = merged_non_anchors(new_ws)

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            if (r, c) in base_skip or (r, c) in new_skip:
                continue
            a = base_ws.cell(row=r, column=c).value
            b = new_ws.cell(row=r, column=c).value

            # Normalize empties
            a_empty = a is None or a == ''
            b_empty = b is None or b == ''

            if a_empty and b_empty:
                continue

            if a_empty and not b_empty:
                yield r, c, 'added'
                continue
            if b_empty and not a_empty:
                yield r, c, 'removed'
                continue

            if not equal_values(a, b, tolerance=tolerance,
                                trim_whitespace=trim_whitespace,
                                case_insensitive=case_insensitive):
                yield r, c, 'changed'


def ensure_unique_sheet_name(wb: Workbook, name: str) -> str:
    if name not in wb.sheetnames:
        return name
    i = 1
    while f"{name}_{i}" in wb.sheetnames:
        i += 1
    return f"{name}_{i}"


def add_summary_sheet(wb_out: Workbook, diffs: List[DiffRecord]) -> None:
    title = ensure_unique_sheet_name(wb_out, "Diff_Summary")
    if title in wb_out.sheetnames:
        ws = wb_out[title]
        wb_out.remove(ws)
    ws = wb_out.create_sheet(title, 0)

    headers = ["Sheet", "Cell", "Status", "Old Value", "New Value"]
    ws.append(headers)

    for d in diffs:
        ws.append([d.sheet, d.cell, d.status, d.old, d.new])

    # Basic formatting: header fill, freeze, autofilter, column widths
    from openpyxl.styles import Font
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(1, len(diffs) + 1)}"
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30


def compare_workbooks(base_path: str,
                      new_path: str,
                      out_path: str,
                      *,
                      include_sheets: Optional[Set[str]] = None,
                      exclude_sheets: Optional[Set[str]] = None,
                      tolerance: float = 0.0,
                      trim_whitespace: bool = False,
                      case_insensitive: bool = False,
                      highlight: bool = True,
                      compare_formulas: bool = False) -> Tuple[int, int]:
    """Compare workbooks and write output.

    Returns: (num_sheets_compared, num_differences)
    """
    # Read values for comparison
    base_wb = load_workbook(base_path, data_only=not compare_formulas, read_only=False)
    new_wb_values = load_workbook(new_path, data_only=not compare_formulas, read_only=False)

    # Load a writable copy of the new workbook for output (preserve formulas/styles)
    new_wb_out = load_workbook(new_path, data_only=False, read_only=False)

    diffs: List[DiffRecord] = []

    base_sheets = set(base_wb.sheetnames)
    new_sheets = set(new_wb_values.sheetnames)

    # Determine sheets to compare
    to_compare = sorted(base_sheets & new_sheets)
    if include_sheets:
        to_compare = [s for s in to_compare if s in include_sheets]
    if exclude_sheets:
        to_compare = [s for s in to_compare if s not in exclude_sheets]

    # Record missing sheets
    only_in_base = base_sheets - new_sheets
    only_in_new = new_sheets - base_sheets
    for s in sorted(only_in_base):
        if include_sheets and s not in include_sheets:
            continue
        if exclude_sheets and s in exclude_sheets:
            continue
        diffs.append(DiffRecord(sheet=s, cell="", status="missing_sheet_new", old=None, new=None))
    for s in sorted(only_in_new):
        if include_sheets and s not in include_sheets:
            continue
        if exclude_sheets and s in exclude_sheets:
            continue
        diffs.append(DiffRecord(sheet=s, cell="", status="missing_sheet_base", old=None, new=None))

    # Compare common sheets
    for s in to_compare:
        base_ws = base_wb[s]
        new_ws_vals = new_wb_values[s]

        # For output highlights, if sheet exists; otherwise skip highlights for this sheet
        new_ws_out = new_wb_out[s] if s in new_wb_out.sheetnames else None

        for r, c, status in iter_diff_cells(base_ws, new_ws_vals,
                                            tolerance=tolerance,
                                            trim_whitespace=trim_whitespace,
                                            case_insensitive=case_insensitive):
            cell = cell_coord(r, c)
            a = base_ws.cell(row=r, column=c).value
            b = new_ws_vals.cell(row=r, column=c).value
            diffs.append(DiffRecord(sheet=s, cell=cell, status=status, old=a, new=b))

            if highlight and new_ws_out is not None:
                out_cell = new_ws_out.cell(row=r, column=c)
                if status == 'added':
                    out_cell.fill = FILL_ADDED
                elif status == 'removed':
                    out_cell.fill = FILL_REMOVED
                elif status == 'changed':
                    out_cell.fill = FILL_CHANGED

    # Add summary sheet to output workbook
    add_summary_sheet(new_wb_out, diffs)

    # Save output
    # Ensure directory exists
    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or '.', exist_ok=True)
    new_wb_out.save(out_path)

    return len(to_compare), len(diffs)


def parse_csv_set(v: Optional[str]) -> Optional[Set[str]]:
    if v is None:
        return None
    items = [x.strip() for x in v.split(',') if x.strip()]
    return set(items) if items else None


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Compare two Excel workbooks across sheets and generate a summary.")
    p.add_argument('--base', '-b', required=True, help='Path to base/original workbook (.xlsx)')
    p.add_argument('--new', '-n', required=True, help='Path to new/updated workbook (.xlsx)')
    p.add_argument('--out', '-o', default=None, help='Path to output diff workbook (.xlsx)')
    p.add_argument('--include-sheets', default=None, help='Comma-separated list of sheet names to include')
    p.add_argument('--exclude-sheets', default=None, help='Comma-separated list of sheet names to exclude')
    p.add_argument('--tolerance', type=float, default=0.0, help='Numeric tolerance for equality (default: 0.0)')
    p.add_argument('--case-insensitive', action='store_true', help='Case-insensitive text comparison')
    p.add_argument('--trim-whitespace', action='store_true', help='Trim whitespace in text comparison')
    p.add_argument('--no-highlights', action='store_true', help='Do not highlight differences in output workbook')
    group = p.add_mutually_exclusive_group()
    group.add_argument('--values', dest='compare_formulas', action='store_false', help='Compare displayed values (default)')
    group.add_argument('--formulas', dest='compare_formulas', action='store_true', help='Compare formulas instead of values')
    p.set_defaults(compare_formulas=False)

    args = p.parse_args(argv)

    base_path = args.base
    new_path = args.new
    out_path = args.out

    if not os.path.isfile(base_path):
        print(f"Base workbook not found: {base_path}", file=sys.stderr)
        return 2
    if not os.path.isfile(new_path):
        print(f"New workbook not found: {new_path}", file=sys.stderr)
        return 2

    if out_path is None:
        root, ext = os.path.splitext(os.path.abspath(new_path))
        out_path = f"{root}_diff{ext or '.xlsx'}"

    include_sheets = parse_csv_set(args.include_sheets)
    exclude_sheets = parse_csv_set(args.exclude_sheets)

    try:
        num_sheets, num_diffs = compare_workbooks(
            base_path=base_path,
            new_path=new_path,
            out_path=out_path,
            include_sheets=include_sheets,
            exclude_sheets=exclude_sheets,
            tolerance=args.tolerance,
            trim_whitespace=args.trim_whitespace,
            case_insensitive=args.case_insensitive,
            highlight=(not args.no_highlights),
            compare_formulas=args.compare_formulas,
        )
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    print(f"Compared {num_sheets} sheet(s). Wrote diff to: {out_path}")
    print(f"Summary entries: {num_diffs}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())

